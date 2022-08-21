import openpyxl

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("sample.xlsx")

# define var to read the sheet
dataframe1 = dataframe.active

# loop to read the dataframe
for row in range(0, dataframe1.max_row):
  for col in dataframe1.iter_cols(1, dataframe1.max_column):
    print(col[row].value)
